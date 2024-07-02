import sys
import requests
from bs4 import BeautifulSoup

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
import os

def write_to_excel(data_dic, file_path="job_info.xlsx"):
    # {--- check if the file was existed ---} #
    file_existed = os.path.exists(file_path)

    if file_existed:
        workbook = load_workbook(file_path)
        sheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        # write table head when create new file #
        for col, key in enumerate(data_dic.keys(), start=1):
            cell = sheet.cell(row=1, column=col, value=key) # write into specified cell
            cell.font = Font(bold=True)
            
    
    # {--- check if the table heads were existed ---} #
    
    # check the next row
    next_row = sheet.max_row + 1
    print(sheet.max_row)

    # check if the next_row == 2 (if yes, the table might be not existed)
    if next_row == 2:
        for col, key in enumerate(data_dic.keys(), start=1):
            cell = sheet.cell(row=1, column=col, value=key)
            cell.font = Font(bold=True)

    # write the data
    for col, key in enumerate(data_dic.keys(), start=1):
        sheet.cell(row=next_row, column=col, value=data_dic[key])

    workbook.save(file_path)

def try_scrap(url):
    response = requests.get(url)
    soup = BeautifulSoup(response.text, "html.parser")
    head_data_dic = {
        "工作名稱": "",
        "公司名稱": "",
        "職務類別": "",
        "工作待遇": "",
        "工作性質": "",
        "上班地點": "",
        "遠端工作": "",
        "上班時段": "",
        "休假制度": "",
        "可上班日": "",
        "需求人數": "",
        "工作經歷": "",
        "學歷要求": "",
        "科系要求": "",
        "語文條件": "",
        "擅長工具": "",
        "工作技能": "",
        "其他條件": "",
    }

    # set the console unicode as "utf-8"
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

    # Get the job title
    job_title = soup.find("div", class_="job-header__title").find("h1").get_text(strip=True)
    head_data_dic["工作名稱"] = job_title

    # Get company
    company = soup.find("div", class_="job-header__title").find("a").get_text(strip=True)
    head_data_dic["公司名稱"] = company

    # Get all the target head and data
    list_row_elements = soup.find_all("div", class_="list-row")
    for element in list_row_elements:
        pick_head_data(head_data_dic, soup, element)

    # execute the excel writing
    write_to_excel(head_data_dic)

    return head_data_dic
    # print(head_data_dic)


# pick the list-row-head and data
def pick_head_data(head_data_dic, soup, element):

    # Get others
    head = element.find("h3").get_text(strip=True)
    data = get_text_from_element(element)
    head_data_dic[head] = data
    

def get_text_from_element(element):
    p_element = element.find("p")
    if p_element:
        return p_element.get_text(strip=True)
    
    div_element = element.find("div", class_="t3 mb-0")
    if div_element:
        return div_element.get_text(strip=True)
    
    return ""
